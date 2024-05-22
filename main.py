import openpyxl
import os

# Открываем существующий файл Excel
file = 'sample.xlsx'
workbook = openpyxl.load_workbook(file)

# Получаем названия всех листов
sheet_names = workbook.sheetnames

# Создаем новую папку для уникальных пользователей
output_folder = "new_folder"
os.makedirs(output_folder, exist_ok=True)

# Обходим все листы
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]

    # Находим номер столбца с заголовком "user_name"
    user_name_column = None
    for col in sheet[1]:
        if col.value == 'user_name':
            user_name_column = col.column
            break

    unique_users = {}

    # Обходим все строки листа, начиная со строки номер 2 (так как первая строка содержит заголовки)
    for row in range(2, sheet.max_row + 1):
        # Значение имени пользователя находится в столбце с заголовком "user_name"
        user_name = sheet.cell(row=row, column=user_name_column).value

        # Если имя пользователя еще не встречалось, добавляем его в словарь
        if user_name not in unique_users:
            unique_users[user_name] = []

        # Добавляем строку с данными этого пользователя в список этого пользователя на текущем листе
        unique_users[user_name].append(tuple(cell.value for cell in sheet[row]))

    # Создаем новые файлы Excel для каждого пользователя на текущем листе
    for user_name, user_data in unique_users.items():
        user_file = f"{output_folder}/{user_name}.xlsx"

        # Если файл для данного пользователя еще не создан, создаем его
        if not os.path.exists(user_file):
            user_workbook = openpyxl.Workbook()
            user_workbook.save(user_file)

        user_workbook = openpyxl.load_workbook(user_file)

        # Создаем лист с данными данного пользователя
        user_sheet = user_workbook.create_sheet(sheet_name)
        user_sheet.append(tuple(cell.value for cell in sheet[1]))

        for row in user_data:
            user_sheet.append(row)

        # Удаляем стандартный пустой лист с названием "Sheet", если он существует перед тем, как сохранить файл

            if "Sheet" in user_workbook.sheetnames:
                default_sheet = user_workbook["Sheet"]
                user_workbook.remove(default_sheet)

        user_workbook.save(user_file)